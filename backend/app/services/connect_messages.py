import csv
import io
from collections.abc import Sequence
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.core.config import settings


CONNECT_MESSAGE_COLUMNS = (
    "Фамилия",
    "Имя",
    "Логин",
    "Статус",
    "Отдел",
    "Профиль доступа",
    "Connect-ссылка",
    "Готовое сообщение",
)
CONNECT_MESSAGE_CSV_COLUMNS = CONNECT_MESSAGE_COLUMNS
CONNECT_MESSAGE_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

CONNECT_MESSAGE_COLUMN_WIDTHS = {
    "A": 20,
    "B": 20,
    "C": 20,
    "D": 15,
    "E": 25,
    "F": 30,
    "G": 70,
    "H": 90,
}


def build_connect_link(user: Any, base_url: str | None = None) -> str:
    token = _clean(getattr(user, "subscription_token", ""))
    if not token:
        return ""
    resolved_base = (base_url if base_url is not None else settings.subscription_base_url).strip().rstrip("/")
    if not resolved_base:
        return f"/connect/{token}"
    return f"{resolved_base}/connect/{token}"


def connect_message_recipient(user: Any) -> str:
    first_name = _clean(getattr(user, "first_name", ""))
    last_name = _clean(getattr(user, "last_name", ""))
    if first_name and last_name:
        return f"{last_name} {first_name}"

    display_name = _clean(getattr(user, "display_name", ""))
    if display_name:
        return display_name

    username = _clean(getattr(user, "username", ""))
    if username:
        return username

    return "пользователь"


def build_connect_message(user: Any, connect_link: str) -> str:
    link = _clean(connect_link) or "Нет connect-ссылки"
    return (
        f"Уважаемый/ая {connect_message_recipient(user)}!\n\n"
        "Это ваша ссылка для подключения к AKFA VPN:\n\n"
        f"{link}\n\n"
        "Перейдите по ссылке и действуйте согласно инструкциям на странице.\n"
        "Не передавайте эту ссылку другим пользователям."
    )


def build_connect_messages_csv(users: Sequence[Any], base_url: str | None = None) -> bytes:
    output = io.StringIO(newline="")
    output.write("\ufeff")
    writer = csv.DictWriter(
        output,
        fieldnames=CONNECT_MESSAGE_COLUMNS,
        delimiter=";",
        lineterminator="\r\n",
    )
    writer.writeheader()
    for user in users:
        writer.writerow(connect_message_row(user, base_url))
    return output.getvalue().encode("utf-8")


def build_connect_messages_xlsx(users: Sequence[Any], base_url: str | None = None) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Сообщения"
    worksheet.append(list(CONNECT_MESSAGE_COLUMNS))

    for user in users:
        row = connect_message_row(user, base_url)
        worksheet.append([row[column] for column in CONNECT_MESSAGE_COLUMNS])
        excel_row = worksheet.max_row
        link_cell = worksheet.cell(excel_row, 7)
        if str(link_cell.value).startswith(("http://", "https://")):
            link_cell.hyperlink = str(link_cell.value)
            link_cell.style = "Hyperlink"

    header_font = Font(bold=True)
    top_alignment = Alignment(vertical="top")
    wrap_top_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = top_alignment

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_top_alignment if cell.column == 8 else top_alignment

    for column, width in CONNECT_MESSAGE_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    if worksheet.max_row >= 1:
        table = Table(displayName="ConnectMessages", ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def connect_message_row(user: Any, base_url: str | None = None) -> dict[str, str]:
    connect_link = build_connect_link(user, base_url)
    department = getattr(user, "department", None)
    profile = getattr(user, "access_profile", None)
    return {
        "Фамилия": _clean(getattr(user, "last_name", "")),
        "Имя": _clean(getattr(user, "first_name", "")),
        "Логин": _clean(getattr(user, "username", "")),
        "Статус": _clean(getattr(user, "status", "")),
        "Отдел": _clean(getattr(department, "name", "")),
        "Профиль доступа": _clean(getattr(profile, "name", "")),
        "Connect-ссылка": connect_link or "Нет connect-ссылки",
        "Готовое сообщение": build_connect_message(user, connect_link),
    }


def _clean(value: object) -> str:
    return str(value or "").strip()
