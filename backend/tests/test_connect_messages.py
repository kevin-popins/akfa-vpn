import csv
import io
from types import SimpleNamespace

from openpyxl import load_workbook

from app.core.config import settings
from app.services.connect_messages import (
    CONNECT_MESSAGE_CSV_COLUMNS,
    CONNECT_MESSAGE_XLSX_CONTENT_TYPE,
    build_connect_message,
)


def test_build_connect_message_uses_last_and_first_name():
    user = SimpleNamespace(first_name="Анна", last_name="Иванова", username="anna.ivanova")
    message = build_connect_message(user, "https://panel.example.test/connect/token")

    assert message.startswith("Уважаемый/ая Иванова Анна!")
    assert "https://panel.example.test/connect/token" in message
    assert "Не передавайте эту ссылку другим пользователям." in message


def test_build_connect_message_fallbacks_without_full_name():
    display_user = SimpleNamespace(first_name="", last_name="", display_name="Мария Петрова", username="maria")
    username_user = SimpleNamespace(first_name="", last_name="", display_name="", username="fallback-login")
    neutral_user = SimpleNamespace(first_name="", last_name="", display_name="", username="")

    assert build_connect_message(display_user, "/connect/a").startswith("Уважаемый/ая Мария Петрова!")
    assert build_connect_message(username_user, "/connect/b").startswith("Уважаемый/ая fallback-login!")
    assert build_connect_message(neutral_user, "/connect/c").startswith("Уважаемый/ая пользователь!")


def test_export_connect_messages_requires_admin_auth(client):
    response = client.get("/admin/users/export-connect-messages")

    assert response.status_code == 401


def test_export_connect_messages_xlsx_contains_columns_link_and_single_message_cell(client, auth_headers):
    old_base_url = settings.subscription_base_url
    settings.subscription_base_url = "https://panel.example.test"
    try:
        created = client.post(
            "/admin/users",
            headers=auth_headers,
            json={"first_name": "Иван", "last_name": "Петров", "username": "ivan.petrov"},
        )
        assert created.status_code == 200
        token = created.json()["subscription_token"]

        response = client.get("/admin/users/export-connect-messages", headers=auth_headers)

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(CONNECT_MESSAGE_XLSX_CONTENT_TYPE)
        assert response.headers["content-disposition"] == 'attachment; filename="akfa-connect-messages.xlsx"'

        connect_link = f"https://panel.example.test/connect/{token}"
        workbook = load_workbook(io.BytesIO(response.content))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        assert headers == list(CONNECT_MESSAGE_CSV_COLUMNS)
        assert worksheet.freeze_panes == "A2"
        assert worksheet.auto_filter.ref == "A1:H2"

        data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 1
        row = data_rows[0]
        assert row[2] == "ivan.petrov"
        assert row[6] == connect_link
        assert row[7].startswith("Уважаемый/ая Петров Иван!")
        assert "\n\nЭто ваша ссылка" in row[7]
        assert connect_link in row[7]
        assert worksheet["H2"].alignment.wrap_text is True
        assert worksheet["G2"].hyperlink.target == connect_link
    finally:
        settings.subscription_base_url = old_base_url


def test_export_connect_messages_csv_fallback_contains_bom_and_semicolon(client, auth_headers):
    created = client.post(
        "/admin/users",
        headers=auth_headers,
        json={"first_name": "Иван", "last_name": "Петров", "username": "ivan.petrov"},
    )
    assert created.status_code == 200

    response = client.get("/admin/users/export-connect-messages?format=csv", headers=auth_headers)

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert response.headers["content-disposition"] == 'attachment; filename="akfa-connect-messages.csv"'
    assert response.content.startswith(b"\xef\xbb\xbf")
    text = response.content.decode("utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(text), delimiter=";"))
    assert list(rows[0].keys()) == list(CONNECT_MESSAGE_CSV_COLUMNS)
    assert rows[0]["Логин"] == "ivan.petrov"


def test_export_connect_messages_can_filter_by_ids(client, auth_headers):
    first = client.post(
        "/admin/users",
        headers=auth_headers,
        json={"first_name": "Олег", "last_name": "Смирнов", "username": "oleg.smirnov"},
    )
    second = client.post(
        "/admin/users",
        headers=auth_headers,
        json={"first_name": "Елена", "last_name": "Орлова", "username": "elena.orlova"},
    )
    assert first.status_code == 200
    assert second.status_code == 200

    response = client.get(
        f"/admin/users/export-connect-messages?ids={first.json()['id']}",
        headers=auth_headers,
    )

    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
    assert [row[2] for row in rows] == ["oleg.smirnov"]
